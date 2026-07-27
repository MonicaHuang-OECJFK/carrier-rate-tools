"""
excel_writer.py（COSCO NEUR 線）
────────────────────────────────
把從 PDF 萃取的 Ocean Freight rate 動態寫入 cheatsheet 的 OFT sheet。

跟 cosco italy 線一樣，用 Mapping tab 做嚴謹比對：
  {(PDF_POL, PDF_POD): (CHEATSHEET_POL, CHEATSHEET_POD)}
不再用「POD 逗號前城市名」這種寬鬆比對 —— PDF 的 POD 只有城市名
（例如 'New York'），跟 cheatsheet 的 'NEW YORK, NY' 對不同城市但
恰好同名的情況（例如兩個不同國家都有 'Savannah'）沒辦法區分，
用 Mapping tab 明確指定對應關係比較安全。

流程：
1. 讀取 Mapping tab → 建立 (PDF_POL, PDF_POD) → (POL, POD) 對照字典
2. 掃描任意 row 找到 POD header → 得到 pod_col 和 header_row
3. POD 右邊 +2 格是 20'，往上驗證 header 是 "20'" → 確認欄位正確
4. +3 = 40'，+4 = 40HC
5. 掃描資料列，用 POL + POD 匹配，寫入 rate
   （cheatsheet 裡用公式連結其他列的 POL，例如 ROTTERDAM/ANTWERP 常常
   用 =H5 這種公式連到 BREMERHAVEN 的數字列 —— 公式格會自動跳過不覆寫，
   也不會被誤報成缺資料）
"""

import openpyxl


def _find_header(ws, keyword):
    """
    掃描整張 sheet，找第一個 value == keyword 的 cell。
    回傳 (row, col) 或拋出 ValueError。
    """
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or "").strip().upper() == keyword.upper():
                return cell.row, cell.column
    raise ValueError(f"找不到 header「{keyword}」")


def _find_oft_cols(ws, pod_col, header_row):
    """
    從 POD 欄往右 +2 格找 OFT 的 20' 欄，
    往上掃 header 驗證那格確實是 '20''。
    確認後 +3 = 40'，+4 = 40HC。

    回傳 (col_20, col_40, col_40hc) 或拋出 ValueError。
    """
    col_20_candidate = pod_col + 2

    confirmed = False
    for r in range(1, header_row + 1):
        val = str(ws.cell(r, col_20_candidate).value or "").strip()
        if val == "20'":
            confirmed = True
            break

    if not confirmed:
        raise ValueError(
            f"POD 右邊 +2 格（col {col_20_candidate}）往上找不到 \"20'\" header，"
            f"請確認 cheatsheet 欄位結構"
        )

    col_20   = col_20_candidate
    col_40   = col_20 + 1
    col_40hc = col_20 + 2

    return col_20, col_40, col_40hc


def _load_mapping(wb):
    """
    讀取 Mapping tab，回傳：
    {(pdf_pol_upper, pdf_pod_upper): (cheatsheet_pol_upper, cheatsheet_pod_upper)}
    """
    if "Mapping" not in wb.sheetnames:
        raise ValueError("找不到 Mapping tab，請確認 cheatsheet 含有 Mapping 分頁")

    ws = wb["Mapping"]
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        pdf_pol, pdf_pod, cs_pol, cs_pod = [str(v or "").strip() for v in row[:4]]
        mapping[(pdf_pol.upper(), pdf_pod.upper())] = (cs_pol.upper(), cs_pod.upper())
    return mapping


def _build_rate_lookup(all_rates, mapping):
    """
    把 PDF 萃取的 all_rates（list of dict，pol/pod 保留 PDF 原始大小寫）
    透過 mapping 轉成 {(POL_upper, POD_upper): (rate_20, rate_40)} 字典。
    """
    lookup = {}
    for row in all_rates:
        pdf_pol = row["pol"].strip().upper()
        pdf_pod = row["pod"].strip().upper()
        key = (pdf_pol, pdf_pod)
        if key in mapping:
            pol, pod = mapping[key]
            lookup[(pol, pod)] = (row["rate_20"], row["rate_40"])
    return lookup


def _write_oft_rates(wb, all_rates, max_scan_rows=95):
    """
    _write_oft_rates 的內部實作，被 update_oft_rates() 跟 update_cheatsheet()
    共用（差別只在誰負責開檔/存檔）。
    """
    ws = wb["OFT"]

    mapping = _load_mapping(wb)
    rate_lookup = _build_rate_lookup(all_rates, mapping)

    pol_header_row, pol_col = _find_header(ws, "POL")
    pod_header_row, pod_col = _find_header(ws, "POD")

    col_20, col_40, col_40hc = _find_oft_cols(ws, pod_col, pod_header_row)

    # 先清空所有 OFT 欄位的值（20' / 40' / 40HC）
    # 只清空純數值的格子，有公式的格子不動
    data_start_row = max(pol_header_row, pod_header_row) + 1
    scan_end_row = data_start_row + max_scan_rows - 1
    for row_num in range(data_start_row, scan_end_row + 1):
        for col in (col_20, col_40, col_40hc):
            cell = ws.cell(row_num, col)
            if cell.data_type != 'f':   # 'f' = formula，公式格不動
                cell.value = None

    updated_count = 0
    skipped_rows = []

    for row_num in range(data_start_row, scan_end_row + 1):
        pol = str(ws.cell(row_num, pol_col).value or "").strip().upper()
        pod = str(ws.cell(row_num, pod_col).value or "").strip().upper()

        if not pol or not pod:
            continue

        cell_20   = ws.cell(row_num, col_20)
        cell_40   = ws.cell(row_num, col_40)
        cell_40hc = ws.cell(row_num, col_40hc)

        # 三格都已經是公式（連結到別列的值）→ 這列本來就不需要直接比對，
        # 直接跳過，不算 skip（避免把「本來就設計成公式」的列誤報成缺資料）
        if cell_20.data_type == 'f' and cell_40.data_type == 'f' and cell_40hc.data_type == 'f':
            continue

        key = (pol, pod)
        if key not in rate_lookup:
            skipped_rows.append((row_num, pol, pod))
            continue

        rate_20, rate_40 = rate_lookup[key]

        if cell_20.data_type != 'f':
            cell_20.value = rate_20
        if cell_40.data_type != 'f':
            cell_40.value = rate_40
        if cell_40hc.data_type != 'f':
            cell_40hc.value = rate_40   # 40HC = 40'

        updated_count += 1

    return updated_count, skipped_rows


def update_oft_rates(excel_path, all_rates, output_path=None, max_scan_rows=95):
    """
    主函式：把 all_rates 寫入 cheatsheet 的 OFT sheet（只更新 OFT，不動 US inland）。

    參數：
      excel_path  : cheatsheet 的路徑（含 Mapping tab）
      all_rates   : cosco_parser.extract_ocean_rates() 產出的 list of dict
                    每筆含 pol / pod / rate_20 / rate_40（pol/pod 為 PDF 原始大小寫）
      output_path : 輸出路徑，None 則直接覆蓋原檔

    回傳：
      (updated_count, skipped_rows)
      updated_count : 成功寫入的列數
      skipped_rows  : 找不到對應 rate 的 (row_num, pol, pod) list
                      （cheatsheet 裡的公式連結列已自動排除，不會出現在這裡）
    """
    wb = openpyxl.load_workbook(excel_path)
    updated_count, skipped_rows = _write_oft_rates(wb, all_rates, max_scan_rows=max_scan_rows)
    out = output_path or excel_path
    wb.save(out)
    return updated_count, skipped_rows


def _write_us_inland(wb, ramp_rows, max_scan_rows=40):
    """
    把 ramp_rows（extract_us_inland 產出）寫入 'US inland' sheet。

    比對用 Location 的城市名（逗號前那段，不分大小寫）——因為 PDF 有時
    會漏寫州名（例如 'Chicago' vs cheatsheet 的 'CHICAGO, IL'）。
    Routing via 也會比對，Location 對但 Routing via 不同的話不寫入，
    另外回報讓你人工確認。

    回傳 (updated_count, skipped_rows, mismatched_rows)
      updated_count   : 成功寫入的列數
      skipped_rows    : cheatsheet 裡找不到對應 Location 的 (row_num, location) list
      mismatched_rows : Location 相符但 Routing via 對不上的
                        (row_num, location, cheatsheet_via, pdf_via) list（不會寫入）
    """
    if "US inland" not in wb.sheetnames:
        raise ValueError("找不到 'US inland' tab")

    ws = wb["US inland"]

    header_row, loc_col = _find_header(ws, "Location")
    _, via_col = _find_header(ws, "Routing via")
    _, col_20 = _find_header(ws, "20DV")
    _, col_40 = _find_header(ws, "40DV/40HQ")
    try:
        _, col_40rf = _find_header(ws, "40RF/40RQ")
    except ValueError:
        col_40rf = None

    def city_of(text):
        return text.strip().upper().split(",")[0].strip()

    ramp_lookup = {city_of(r["location"]): r for r in ramp_rows}

    data_start_row = header_row + 1
    scan_end_row = data_start_row + max_scan_rows - 1

    updated_count = 0
    skipped_rows = []
    mismatched_rows = []

    for row_num in range(data_start_row, scan_end_row + 1):
        location = str(ws.cell(row_num, loc_col).value or "").strip()
        if not location:
            continue

        city = city_of(location)
        if city not in ramp_lookup:
            skipped_rows.append((row_num, location))
            continue

        ramp = ramp_lookup[city]

        cheatsheet_via = str(ws.cell(row_num, via_col).value or "").strip().upper()
        pdf_via = ramp["routing_via"].strip().upper()
        if cheatsheet_via and pdf_via and cheatsheet_via != pdf_via:
            mismatched_rows.append((row_num, location, cheatsheet_via, pdf_via))
            continue

        cell_20 = ws.cell(row_num, col_20)
        if cell_20.data_type != 'f':
            cell_20.value = ramp["rate_20"] if ramp["rate_20"] is not None else "-"

        cell_40 = ws.cell(row_num, col_40)
        if cell_40.data_type != 'f':
            cell_40.value = ramp["rate_40"] if ramp["rate_40"] is not None else "-"

        if col_40rf is not None and ramp.get("rate_40rf") is not None:
            cell_40rf = ws.cell(row_num, col_40rf)
            if cell_40rf.data_type != 'f':
                cell_40rf.value = ramp["rate_40rf"]

        updated_count += 1

    return updated_count, skipped_rows, mismatched_rows


def update_us_inland_only(excel_path, ramp_rows, output_path=None):
    """
    主函式：只把 ramp_rows 寫入 cheatsheet 的 'US inland' sheet，不動 OFT。

    COSCO NEUR 這條線的 US inland 費率有時候會跟 Ocean Freight 分開寄一份
    PDF，這個函式讓你可以單獨跑 US inland 更新，不需要同時準備 Ocean
    Freight 的 PDF。

    參數：
      excel_path  : cheatsheet 的路徑
      ramp_rows   : cosco_parser.extract_us_inland() 產出的 list of dict
      output_path : 輸出路徑，None 則直接覆蓋原檔

    回傳：
      (updated_count, skipped_rows, mismatched_rows) —— 跟 _write_us_inland 一樣
    """
    wb = openpyxl.load_workbook(excel_path)
    updated_count, skipped_rows, mismatched_rows = _write_us_inland(wb, ramp_rows)
    out = output_path or excel_path
    wb.save(out)
    return updated_count, skipped_rows, mismatched_rows


def update_cheatsheet(excel_path, all_rates, ramp_rows=None, output_path=None, max_scan_rows=95):
    """
    主函式：把 PDF 萃取結果寫入 cheatsheet（單次讀檔、單次存檔）。

    參數：
      excel_path  : cheatsheet 的路徑（含 Mapping tab）
      all_rates   : cosco_parser.extract_ocean_rates() 產出的 list of dict
      ramp_rows   : cosco_parser.extract_us_inland() 產出的 list of dict。
                    這張表不是每期 PDF 都有 —— 傳 None 或空 list 就完全不動
                    US inland tab（不會報錯，也不會清空原本的值）
      output_path : 輸出路徑，None 則直接覆蓋原檔

    回傳 dict：
      {
        "oft_updated":       int,
        "oft_skipped":       [(row_num, pol, pod), ...],
        "inland_updated":    int,
        "inland_skipped":    [(row_num, location), ...],
        "inland_mismatched": [(row_num, location, cheatsheet_via, pdf_via), ...],
      }
    """
    wb = openpyxl.load_workbook(excel_path)

    oft_updated, oft_skipped = _write_oft_rates(wb, all_rates, max_scan_rows=max_scan_rows)

    inland_updated, inland_skipped, inland_mismatched = 0, [], []
    if ramp_rows:
        inland_updated, inland_skipped, inland_mismatched = _write_us_inland(wb, ramp_rows)

    out = output_path or excel_path
    wb.save(out)

    return {
        "oft_updated":       oft_updated,
        "oft_skipped":       oft_skipped,
        "inland_updated":    inland_updated,
        "inland_skipped":    inland_skipped,
        "inland_mismatched": inland_mismatched,
    }
