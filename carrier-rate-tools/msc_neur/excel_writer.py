"""
excel_writer.py (MSC 版)
────────────────────────
讀取 cheatsheet 的 Mapping tab，把從 MSC PDF 萃取的
20DV / 40DV(HC) 運費寫入 rate sheet 的欄位。

跟 EMC 那份不同：
  - PDF 萃取出來的 POL/POD（如 "Antwerp" / "Baltimore"）
    跟 cheatsheet 裡的 POL/POD（如 "ANTWERP" / "BALTIMORE, MD"）
    寫法不一樣，所以需要 Mapping tab 做轉換
    （格式跟 COSCO 那份類似：pdf_pol, pdf_pod, cs_pol, cs_pod 四欄。
    注意 cheatsheet 裡還有一欄 POR，那是起運地區代碼，不是拿來比對
    PDF 港口名稱用的，比對要用 POL 欄）
  - 欄位結構：POD 往右 +2 = 20'，+3 = 40'，+4 = 40HC
  - cheatsheet 裡很多列是公式（例如 Rotterdam 那幾列引用 Antwerp
    的數字），這些公式格「一律不覆寫」，只填純數字的格子。
    判斷方式：cell.data_type == 'f' 就跳過。

流程：
1. 讀取 Mapping tab → 建立 (PDF_POL, PDF_POD) → (cheatsheet POL, POD) 對照字典
2. 掃描整張 sheet 找 POL / POD header → 得到 col + header_row
3. POD 右邊 +2 格是 20'，往上驗證 header 是 "20'" → 確認欄位正確
4. +3 = 40'，+4 = 40HC
5. 掃描資料列，用 (POL, POD) 比對 rate_lookup，命中且不是公式格才寫入

依賴：pip install openpyxl
"""

import openpyxl


def _find_header(ws, keyword):
    """掃描整張 sheet，找第一個 value == keyword 的 cell。回傳 (row, col)。"""
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or "").strip().upper() == keyword.upper():
                return cell.row, cell.column
    raise ValueError(f"找不到 header「{keyword}」")


def _find_rate_cols(ws, pod_col, header_row):
    """
    從 POD 欄往右 +2 找 20' 欄，往上掃這欄驗證有沒有 "20'" header。
    確認後 +3 = 40'，+4 = 40HC。

    回傳 (col_20, col_40, col_40hc)。
    """
    col_20_candidate = pod_col + 2

    confirmed = False
    for r in range(1, header_row + 1):
        val = str(ws.cell(r, col_20_candidate).value or "").strip()
        if val == "20'":
            confirmed = True
            break

    if not confirmed:
        raise ValueError(
            f"POD 右邊 +2 格（col {col_20_candidate}）往上找不到 \"20'\" header，"
            f"請確認 cheatsheet 欄位結構是否改動"
        )

    col_20   = col_20_candidate
    col_40   = col_20 + 1
    col_40hc = col_20 + 2

    return col_20, col_40, col_40hc


def _load_mapping(wb):
    """
    讀取 Mapping tab，回傳：
    {(pdf_pol_upper, pdf_pod_upper): (cheatsheet_pol_upper, cheatsheet_pod_upper)}
    """
    if "Mapping" not in wb.sheetnames:
        raise ValueError("找不到 Mapping tab，請確認 cheatsheet 含有 Mapping 分頁")

    ws = wb["Mapping"]
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        pdf_pol, pdf_pod, cs_pol, cs_pod = [str(v or "").strip() for v in row[:4]]
        mapping[(pdf_pol.upper(), pdf_pod.upper())] = (cs_pol.upper(), cs_pod.upper())
    return mapping


def _build_rate_lookup(all_rates, mapping):
    """
    把 PDF 萃取的 all_rates（list of dict，含 pol/pod/rate_20/rate_40）
    透過 mapping 轉成 {(POL_upper, POD_upper): (rate_20, rate_40)} 字典。
    """
    lookup = {}
    for row in all_rates:
        pdf_pol = row["pol"].strip().upper()
        pdf_pod = row["pod"].strip().upper()
        key = (pdf_pol, pdf_pod)
        if key in mapping:
            pol, pod = mapping[key]
            lookup[(pol, pod)] = (row["rate_20"], row["rate_40"])
    return lookup


def update_rates(excel_path, all_rates, output_path=None, sheet_name="Sheet1", max_scan_rows=200):
    """
    主函式：把 all_rates 寫入 cheatsheet。

    參數：
      excel_path  : cheatsheet 路徑（含 Mapping tab）
      all_rates   : msc_extract.extract() 產出的 list of dict
                    每筆含 pol / pod / rate_20 / rate_40
      output_path : 輸出路徑，None 則直接覆蓋原檔
      sheet_name  : 要寫入的分頁名稱

    回傳：
      (updated_count, skipped_rows, formula_skipped)
      updated_count    : 成功寫入的列數（至少一欄被寫入）
      skipped_rows     : 找不到對應 rate 的 (row_num, pol, pod) list
      formula_skipped   : 有對應 rate 但因為是公式格而跳過的 (row_num, pol, pod) list
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]

    mapping = _load_mapping(wb)
    rate_lookup = _build_rate_lookup(all_rates, mapping)

    pol_header_row, pol_col = _find_header(ws, "POL")
    pod_header_row, pod_col = _find_header(ws, "POD")
    col_20, col_40, col_40hc = _find_rate_cols(ws, pod_col, pod_header_row)

    data_start_row = max(pol_header_row, pod_header_row) + 1
    scan_end_row = data_start_row + max_scan_rows - 1

    updated_count = 0
    skipped_rows = []
    formula_skipped = []

    for row_num in range(data_start_row, scan_end_row + 1):
        pol = str(ws.cell(row_num, pol_col).value or "").strip().upper()
        pod = str(ws.cell(row_num, pod_col).value or "").strip().upper()

        if not pol and not pod:
            continue

        key = (pol, pod)
        if key not in rate_lookup:
            skipped_rows.append((row_num, pol, pod))
            continue

        rate_20, rate_40 = rate_lookup[key]
        wrote_any = False
        all_formula = True

        for col, val in ((col_20, rate_20), (col_40, rate_40), (col_40hc, rate_40)):
            cell = ws.cell(row_num, col)
            if cell.data_type == 'f':   # 公式格，跳過不覆寫
                continue
            all_formula = False
            cell.value = val
            wrote_any = True

        if wrote_any:
            updated_count += 1
        elif all_formula:
            formula_skipped.append((row_num, pol, pod))

    out = output_path or excel_path
    wb.save(out)
    return updated_count, skipped_rows, formula_skipped
